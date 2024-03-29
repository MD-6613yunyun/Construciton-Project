from flask import Blueprint,render_template, request, redirect , url_for , jsonify, session
from website import db_connect,catch_db_insert_error
from openpyxl import load_workbook
from psycopg2 import IntegrityError
import calendar
from datetime import datetime

views = Blueprint('views',__name__)

machine_type_mapping = {}
machine_class_mapping = {}
machine_owner_mapping = {}
business_unit_mapping = {}
machine_brand_mapping = {}
machine_capacity_mapping = {}
mappings = []
day_max = {'1':31,'2':31,'3':31,'4':30,'5':31,'6':30,'7':31,'8':31,'9':30,'10':31,'11':30,'12':31}


def get_first_and_last_day(month, year):
    _, last_day = calendar.monthrange(year, month)
    first_day = f"{year}-{month:02d}-01"
    last_day = f"{year}-{month:02d}-{last_day}"
    return first_day, last_day


@views.route('/')
def home():
    conn = db_connect()
    cur = conn.cursor()
    if "cpu_user_id" not in session:
        return redirect(url_for("auth.authenticate"))
    user_id = session["cpu_user_id"]
    cur.execute("SELECT user_access_id FROM user_auth WHERE id = %s;",(user_id,))
    current_role = cur.fetchone()[0]
    if user_id and current_role:
        if current_role in [3,4]:
            cur.execute("SELECT pj.id,pj.code,pj.name FROM analytic_project_code AS pj;")
        else:
            cur.execute("SELECT pj.id,pj.code,pj.name FROM project_user_access access LEFT JOIN  analytic_project_code pj ON access.project_id = pj.id WHERE access.user_id = %s;",(user_id,))
        project_datas = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('home.html',project_datas = project_datas, current_role = current_role)
    return render_template('auth.html',typ='log')

@views.route("/transactions/<what>/<mgs>",methods =['GET','POST'])
@views.route("/transactions/<what>",methods =['GET','POST'])
def show_transactions(what,mgs=None):

    if "cpu_user_id" not in session:
        return redirect(url_for("auth.authenticate"))

    conn = db_connect()
    cur = conn.cursor()

    user_id = session["cpu_user_id"]
    cur.execute("SELECT user_access_id FROM user_auth WHERE id = %s;",(user_id,))
    role = cur.fetchone()[0]

    search_wildcard = ""

    if not role:
        return redirect(url_for('views.home'))
    else:
        if role not in (1,3,4):
            return render_template('access_error.html')
    if request.method == 'POST':
        if what == 'search':
            search_value = request.form.get('search-value')
            for_what = request.form.get('for-what').strip()
            # print(for_what)
            what_dct = {'Income Expense Query':'income-expense','Duty Query':'duty','Expenses Query':'expense'}
            what = what_dct.get(for_what,for_what)
            search_wildcard = search_value.strip()
            # print(what_dct)
            # print(what)
        else:
            return render_template('access_error.html')

    if role in (3, 4):
        cur.execute("SELECT pj.id,pj.code,pj.name FROM analytic_project_code AS pj;")
    else:
        cur.execute("SELECT pj.id,pj.code,pj.name FROM project_user_access access LEFT JOIN  analytic_project_code pj ON access.project_id = pj.id WHERE access.user_id = %s;",(user_id,))
    project_datas = cur.fetchall()
    if what == 'duty':
        query = """ SELECT 
                        dty.duty_date,pj.code,pj.name,fv.machine_name,dty.operator_name,dty.morg_start,dty.morg_end,
                        dty.aftn_start,dty.aftn_end,dty.evn_start,dty.evn_end,dty.total_hr,dty.hrper_rate,
                        dty.totaluse_fuel,dty.fuel_price,dty.duty_amt,dty.fuel_amt,dty.total_amt,dty.way,
                        dty.complete_feet, dty.complete_sud 
                    FROM duty_odoo_report dty
                        LEFT JOIN fleet_vehicle fv ON fv.id = dty.machine_id
                        LEFT JOIN analytic_project_code pj ON pj.id = dty.project_id ORDER BY dty.duty_date ASC LIMIT 81"""
        cur.execute(query)
        datas = cur.fetchall()
        cur.execute("SELECT count(*) FROM duty_odoo_report;")
        total = cur.fetchall()[0][0]
        name = "Duty Query"
    elif what == 'expense':
        query = """ SELECT
                        exp.duty_date,pj.code,pj.name,bi.name,exp.expense_amt
                    FROM expense_prepaid AS exp
                    INNER JOIN analytic_project_code AS pj
                    ON pj.id = exp.project_id
                    INNER JOIN res_company bi
                    ON bi.id = exp.res_company_id
                    LIMIT 81; """
        cur.execute(query)
        datas = cur.fetchall()
        cur.execute("SELECT count(*) FROM expense_prepaid;")
        total = cur.fetchall()[0][0]
        name = "Expenses Query"
    elif what == 'income-expense':
        cur.execute(""" 
            SELECT form.set_date,form.income_expense_no,pj.code,pj.name,
            CASE WHEN form.income_status = 't' THEN 'Income' ELSE 'Expense' END,
            line.description,car.machine_name,line.invoice_no,line.qty,line.price,line.amt,line.remark 
            FROM income_expense_line line 
            LEFT JOIN income_expense form 
            ON line.income_expense_id  = form.id
            LEFT JOIN analytic_project_code pj
            ON pj.id = form.project_id
            LEFT JOIN fleet_vehicle car
            ON car.id = line.machine_id
            WHERE pj.code ILIKE %s OR pj.name ILIKE %s
            ORDER BY 
                CASE 
                    WHEN pj.name = %s THEN 0 WHEN pj.code = %s 
                    THEN 0 ELSE 1 
                END,form.project_id,form.set_date DESC;""",('%'+search_wildcard+'%','%'+search_wildcard+'%',search_wildcard,search_wildcard))
        datas = cur.fetchall()
        cur.execute("SELECT count(id) FROM income_expense_line;")
        total = cur.fetchall()[0][0]
        name = "Income Expense Query"
    elif what == 'machine-activity':
        cur.execute(""" SELECT form.set_date,pj.code,pj.name,form.daily_activity_no,car.machine_name,ajt.name,ajf.name,line.duty_hour,line.used_fuel,line.description,COALESCE(line.way,0.0) FROM daily_activity_lines AS line 
                    LEFT JOIN daily_activity AS form ON line.daily_activity_id = form.id 
                    LEFT JOIN analytic_project_code AS pj ON pj.id = form.project_id 
                    LEFT JOIN fleet_vehicle AS car ON car.id = line.machine_id 
                    LEFT JOIN activity_job_type AS ajt ON ajt.id = line.job_type_id 
                    LEFT JOIN activity_job_function AS ajf ON ajf.id = line.job_function_id 
                    ORDER BY pj.id,form.set_date;""")
        datas = cur.fetchall()
        cur.execute("SELECT count(*) FROM daily_activity_lines;")
        total = cur.fetchone()[0]
        name = "Machine Activities Query"
    elif what == 'repair-activity':
        cur.execute(""" SELECT form.set_date,pj.code,pj.name,form.daily_activity_no,car.machine_name,line.description,line.accident_status 
                    FROM daily_activity_accident_lines AS line 
                    LEFT JOIN daily_activity AS form ON line.daily_activity_id = form.id 
                    LEFT JOIN analytic_project_code AS pj ON pj.id = form.project_id 
                    LEFT JOIN fleet_vehicle AS car ON car.id = line.machine_id
                    ORDER BY pj.id,form.set_date; """)
        datas = cur.fetchall()
        cur.execute("SELECT count(*) FROM daily_activity_accident_lines;")
        total = cur.fetchone()[0]
        name = "Repair Activities Query"
    elif what == 'service':
        datas = []
        name = "Service Query"
        total = 70
    return render_template("transactions.html",datas=datas,total=total,name=name,mgs=mgs,project_datas = project_datas, current_role = role)

@views.route("/configurations/<what>/<mgs>",methods=['GET','POST'])
@views.route("/configurations/<what>",methods=['GET','POST'])
def configurations(what,mgs=None):

    if "cpu_user_id" not in session:
        return redirect(url_for("auth.authenticate"))    

    conn = db_connect()
    cur = conn.cursor()

    user_id = session["cpu_user_id"]
    cur.execute("SELECT user_access_id FROM user_auth WHERE id = %s;",(user_id,))
    role = cur.fetchone()[0]

    if not role:
        return redirect(url_for('views.home'))
    else:
        if role not in (1,3,4):
            return render_template('access_error.html')
    if role in (3,4):
        cur.execute("SELECT pj.id,pj.code,pj.name FROM analytic_project_code AS pj;")
    else:
        cur.execute("SELECT pj.id,pj.code,pj.name FROM project_user_access access LEFT JOIN  analytic_project_code pj ON access.project_id = pj.id WHERE access.user_id = %s;",(user_id,))
    project_datas = cur.fetchall()

    search_wildcard = ""
    where_filter_clause = "         "
    extra_datas = ["",False,0]
    project_stat_form_id = None
    if request.method == 'POST':
        if what == 'search':
            search_value = request.form.get('search-value')
            for_what = request.form.get('for-what')
            what_dct = {'Machine List':'machine','Project List':'project','Project Statistics':'project-stat'}
            what = what_dct.get(for_what,for_what)
            where_filter_clause = "WHERE"
            if what == 'project':
                pj_name = request.form.get('pj-name').strip()
                if pj_name != '':
                    where_filter_clause += f" ( pj.code ilike '%{pj_name}%' or pj.name ilike '%{pj_name}%' ) AND   "
                pj_group = request.form.get('pj-group').strip()
                if pj_group != '':
                    where_filter_clause += f" p_group.name ilike '%{pj_group}%' AND   "
                pj_type = request.form.get('pj-type').strip()
                if pj_type != '':
                    where_filter_clause += f" p_type.name ilike '%{pj_type}%' AND   "
                unit = request.form.get('unit').strip()
                if unit != '':
                    where_filter_clause += f" bi.name ilike '%{unit}%' AND   "
            elif what == 'project-stat':
                pj_name = request.form.get('pj-name').strip()
                if pj_name != '':
                    where_filter_clause += f" ( pj.code ilike '%{pj_name}%' or pj.name ilike '%{pj_name}%' ) AND   "
                emp = request.form.get('supervisor').strip()
                if emp != '':
                    where_filter_clause += f" emp.name ilike '%{emp}%' AND   "
                loc = request.form.get('location').strip()
                if loc != '':
                    where_filter_clause += f" stat.location ilike '%{loc}%' AND   "
                start_dt = request.form.get('start-dt').strip()
                if start_dt != '':
                    where_filter_clause += f" stat.pj_start_date ilike '%{start_dt}%' AND   "                    
            elif what == 'machine':
                pj_name = request.form.get('pj-name').strip()
                if pj_name != '':
                    where_filter_clause += f" ( pj.code ilike '%{pj_name}%' or pj.name ilike '%{pj_name}%' ) AND   "
                machine_name = request.form.get("name").strip()
                if machine_name != '':
                    where_filter_clause += f" fv.machine_name ilike '%{machine_name}%' AND   "
                machine_type = request.form.get("type").strip()
                if machine_type != '':
                    where_filter_clause += f" mt.name ilike '%{machine_type}%' AND   "
                machine_class = request.form.get("class").strip()
                if machine_class != '':
                    where_filter_clause += f" mc.name ilike '%{machine_class}%' AND   "
                unit  = request.form.get("unit").strip()
                if unit != '':
                    where_filter_clause += f" rc.name ilike '%{unit}%' AND   "
                capacity = request.form.get("capacity").strip()
                if capacity != '':
                    where_filter_clause += f" mcf.name ilike '%{capacity}%' AND   "
                brand = request.form.get("brand").strip()
                if brand != '':
                    where_filter_clause += f" vb.name ilike '%{brand}%' AND   "
                owner = request.form.get("owner").strip()
                if owner != '':
                    where_filter_clause += f" vo.name ilike '%{owner}%' AND   "
            search_wildcard = search_value
            if search_wildcard != "":
                extra_datas[1] = True
        else:
            edit_id = request.form.get("edit-id")
            datalist_datas = {}
            if what == 'Machine List':
                price_type_id = request.form.get("price_type_id")
                project_stat_form_id = request.form.get("stat-form-id")
                if price_type_id:
                    duty_price_edit_id = request.form.get("duty-price-edit-id")
                    # print(duty_price_edit_id)
                    machine_type_id = request.form.get("machine_type_id")
                    start_date = request.form.get("start_date")
                    start_date = datetime.strptime(start_date, "%d/%m/%Y").date()
                    end_date = request.form.get("end_date")             
                    end_date = None if end_date.strip() == "" else datetime.strptime(end_date, "%d/%m/%Y").date()
                    price = request.form.get("price")
                    duty_price_history_id = duty_price_edit_id if duty_price_edit_id else 0
                    if end_date:
                        cur.execute("""  
                                    SELECT COUNT(*)
                                        FROM duty_price_history
                                    WHERE 
                                        machine_id = %s AND id <> %s AND
                                        (
                                            (start_date <= %s AND end_date >= %s) OR (start_date <= %s AND end_date >= %s) OR
                                            (start_date >= %s AND end_date <= %s)
                                        )
                                    ;
                                """,(edit_id,duty_price_history_id,start_date,start_date,end_date,end_date,start_date,end_date))
                    else:
                        cur.execute(""" 
                            SELECT COUNT(*) FROM duty_price_history
                                WHERE 
                                    machine_id = %s AND id <> %s AND
                                    (
                                        (%s BETWEEN start_date AND  end_date) 
                                        OR  (%s >= start_date AND %s < COALESCE(end_date,%s::date-1)) 
                                        OR (start_date = %s)
                                    )
                                    ;
                        """,(edit_id,duty_price_history_id,start_date,start_date,start_date,start_date,start_date))
                    duplicate = cur.fetchone()[0]
                    if duplicate != 0 or (end_date and end_date < start_date):
                        mgs = "Invalid Dates!! Please Double Check Dates .."
                    cur.execute("SELECT id FROM duty_price_history WHERE machine_id = %s AND machine_type_id = %s AND duty_price_type_id = %s AND duty_price = %s AND start_date = %s AND end_date = %s;",(edit_id,machine_type_id,price_type_id,price,start_date,end_date))
                    if not cur.fetchone() and mgs is None:
                        if duty_price_edit_id:
                            cur.execute("SELECT start_date,COALESCE(end_date,start_date) FROM duty_price_history WHERE id = %s;",(duty_price_edit_id,))
                            origin_dates = cur.fetchone()
                            if start_date > origin_dates[0] or (end_date and end_date < origin_dates[1]):
                                mgs = "You can't shrink dates while deleting!!"
                            else:
                                cur.execute("UPDATE duty_price_history SET machine_type_id = %s, duty_price_type_id = %s , duty_price = %s , start_date = %s , end_date = %s WHERE id = %s;",(machine_type_id, price_type_id, price, start_date, end_date, duty_price_edit_id))
                        else:
                            cur.execute("SELECT MIN(start_date) FROM duty_price_history WHERE machine_id = %s UNION ( SELECT start_date FROM duty_price_history WHERE machine_id = %s ORDER BY start_date);",(edit_id,edit_id))
                            datas = cur.fetchall()
                            # print(start_date)
                            # print(datas)
                            if not datas[0][0]:
                                cur.execute("INSERT INTO duty_price_history (machine_id,machine_type_id,duty_price_type_id,duty_price,start_date,end_date) VALUES (%s,%s,%s,%s,%s,%s);",(edit_id,machine_type_id,price_type_id,price,start_date,end_date))                
                            elif start_date < datas[0][0]:
                                if end_date:
                                    cur.execute("INSERT INTO duty_price_history (machine_id,machine_type_id,duty_price_type_id,duty_price,start_date,end_date) VALUES (%s,%s,%s,%s,%s,%s);",(edit_id,machine_type_id,price_type_id,price,start_date,end_date))
                                else:
                                    cur.execute("INSERT INTO duty_price_history (machine_id,machine_type_id,duty_price_type_id,duty_price,start_date,end_date) VALUES (%s,%s,%s,%s,%s,(SELECT MIN(start_date) FROM duty_price_history WHERE machine_id = %s)- INTERVAL '1 day');",(edit_id,machine_type_id,price_type_id,price,start_date, edit_id))
                            else:
                                if start_date > datas[-1][0]:
                                    cur.execute("UPDATE duty_price_history SET end_date = (SELECT %s::date -1) WHERE machine_id = %s AND end_date IS NULL;",(start_date,edit_id))
                                cur.execute("INSERT INTO duty_price_history (machine_id,machine_type_id,duty_price_type_id,duty_price,start_date,end_date) VALUES (%s,%s,%s,%s,%s,%s);",(edit_id,machine_type_id,price_type_id,price,start_date,end_date))
                        conn.commit()
                    else:
                        mgs = "Same Datas are existed!!!"
                cur.execute("SELECT name,id FROM machine_type")
                datas = cur.fetchall()
                datalist_datas["Machine Type"] = datas

                cur.execute("SELECT name FROM machine_class")
                datas = cur.fetchall()
                datalist_datas["Machine Class"] = datas

                cur.execute("SELECT name FROM res_company")
                datas = cur.fetchall()
                datalist_datas["Business Unit"] = datas

                cur.execute("SELECT name FROM vehicle_machine_config")
                datas = cur.fetchall()
                datalist_datas["Machine Capacity"] = datas

                cur.execute("SELECT name FROM fleet_vehicle_model_brand")
                datas = cur.fetchall()
                datalist_datas["Machine Brand"] = datas

                cur.execute("SELECT name FROM vehicle_owner")
                datas = cur.fetchall()
                datalist_datas["Owner"] = datas

                cur.execute("""SELECT his.id,mt.name,dtyType.name,his.start_date,his.end_date,his.duty_price,mt.id,dtyType.id 
                                FROM duty_price_history AS his
                                LEFT JOIN machine_type mt ON mt.id = his.machine_type_id
                                LEFT JOIN duty_price_type dtyType ON dtyType.id = his.duty_price_type_id
                                WHERE his.machine_id = %s
                                ORDER BY start_date DESC;""",(edit_id,))
                datas = cur.fetchall()
                datalist_datas['history'] = datas


                cur.execute("SELECT id,name FROM duty_price_type;")
                datalist_datas['price_type'] = cur.fetchall()

                cur.execute("""SELECT fv.machine_name,mt.name,mc.name,rc.name,mcf.name,vb.name,vo.name,fv.id
                    FROM fleet_vehicle fv
                    LEFT JOIN machine_type mt ON mt.id = fv.machine_type_id 
                    LEFT JOIN machine_class mc ON mc.id = fv.machine_class_id
                    LEFT JOIN res_company rc ON rc.id = fv.business_unit_id
                    LEFT JOIN fleet_vehicle_model_brand vb ON vb.id = fv.brand_id
                    LEFT JOIN vehicle_owner vo ON vo.id = fv.owner_name_id
                    LEFT JOIN vehicle_machine_config mcf ON mcf.id = fv.machine_config_id
                    WHERE fv.id = %s;""",(edit_id,))
                datalist_datas['name'] = 'Machine Edit'
            elif what == 'Project List':
                cur.execute("SELECT name FROM res_company;")
                datalist_datas["Business Unit"] = cur.fetchall()

                cur.execute("SELECT name FROM project_group;")
                datalist_datas["Project Group"] = cur.fetchall()

                cur.execute("SELECT name FROM project_type;")
                datalist_datas["Project Type"] = cur.fetchall()

                datalist_datas['history'] = []

                cur.execute("""
                    SELECT DISTINCT ON (car.id,his.start_time)
                        his.id,
                        car.machine_name,
                        his.start_time::date AS project_start_date,
                        type.name,
                        price_his.duty_price AS duty_price
                    FROM
                        machines_history AS his
                    INNER JOIN
                        fleet_vehicle AS car ON his.machine_id = car.id
                    LEFT JOIN
                        duty_price_history AS price_his ON price_his.machine_id = his.machine_id 
                    LEFT JOIN 
	                    duty_price_type AS type ON type.id = price_his.duty_price_type_id
                    WHERE
                        his.project_id = %s AND his.end_time IS  NULL 
                            AND (
                            his.start_time::date BETWEEN price_his.start_date AND COALESCE(price_his.end_date,CURRENT_DATE)
                            OR his.end_time::date BETWEEN price_his.start_date AND COALESCE(price_his.end_date,CURRENT_DATE)
                        ) ORDER BY his.start_time DESC;
                """,(edit_id,))
                datalist_datas['history'].append(cur.fetchall())

                cur.execute("""
                    SELECT DISTINCT ON (car.id,his.start_time)
                        his.id,
                        car.machine_name,
                        his.start_time::date AS project_start_date,
                        his.end_time::date AS project_end_date,
                        type.name,
                        price_his.duty_price AS duty_price
                    FROM
                        machines_history AS his
                    INNER JOIN
                        fleet_vehicle AS car ON his.machine_id = car.id
                    LEFT JOIN
                        duty_price_history AS price_his ON price_his.machine_id = his.machine_id 
                    LEFT JOIN 
	                    duty_price_type AS type ON type.id = price_his.duty_price_type_id
                    WHERE
                        his.project_id = %s AND his.end_time IS NOT NULL 
                            AND (
                            his.start_time::date BETWEEN price_his.start_date AND COALESCE(price_his.end_date,CURRENT_DATE)
                            OR his.end_time::date BETWEEN price_his.start_date AND COALESCE(price_his.end_date,CURRENT_DATE)
                        ) ORDER BY his.start_time DESC;
                """,(edit_id,))
                datalist_datas['history'].append(cur.fetchall())

                cur.execute(""" SELECT 
                            code,pj.name,p_group.name,pj_type.name,bi.name,pj.id
                            FROM analytic_project_code AS pj 
                            LEFT JOIN res_company AS bi
                            ON pj.business_unit_id = bi.id
                            LEFT JOIN project_type AS pj_type
                            ON pj_type.id = pj.pj_type_id
                            LEFT JOIN project_group AS p_group  
                            ON p_group.id = pj.pj_group_id
                            WHERE pj.id = %s;""",(edit_id,))
                datalist_datas['name'] = 'Project Edit'
            data = cur.fetchone()
            # print(data)
            # print(project_stat_form_id)
            return render_template('edit_configurations.html',mgs=mgs,data=data,what=what,project_datas=project_datas,datalist_datas=datalist_datas,project_stat_form_id=project_stat_form_id, current_role = role)
            

    if what == 'machine':
        cur.execute(f"""SELECT pj.name,fv.machine_name,mt.name,mc.name,rc.name,mcf.name,vb.name,vo.name,fv.id
                    FROM fleet_vehicle fv
                    LEFT JOIN machine_type mt ON mt.id = fv.machine_type_id 
                    LEFT JOIN machine_class mc ON mc.id = fv.machine_class_id
                    LEFT JOIN res_company rc ON rc.id = fv.business_unit_id
                    LEFT JOIN fleet_vehicle_model_brand vb ON vb.id = fv.brand_id
                    LEFT JOIN vehicle_owner vo ON vo.id = fv.owner_name_id
                    LEFT JOIN vehicle_machine_config mcf ON mcf.id = fv.machine_config_id
                    LEFT JOIN 
                        ( SELECT project_id,machine_id FROM machines_history 
                            WHERE CURRENT_TIMESTAMP BETWEEN start_time AND COALESCE(end_time,CURRENT_TIMESTAMP)
                        ) AS history 
                        ON history.machine_id = fv.id
                    LEFT JOIN analytic_project_code AS pj
                        ON pj.id = history.project_id
                    {where_filter_clause[:-6]}
                    LIMIT 81;""")
        data = cur.fetchall()
        cur.execute("SELECT count(id) FROM fleet_vehicle;")
        extra_datas[2] = cur.fetchall()[0][0]
        extra_datas[0] = "Machine List"
    elif what == 'project':
        cur.execute(f"""SELECT 
                    code,pj.name,p_group.name,p_type.name,bi.name,pj.id
                    FROM analytic_project_code pj 
                    LEFT JOIN res_company bi
                    ON pj.business_unit_id = bi.id
                    LEFT JOIN project_group AS p_group
                    ON p_group.id = pj.pj_group_id
                    LEFT JOIN project_type AS p_type
                    ON p_type.id = pj.pj_type_id
                    {where_filter_clause[:-6]}
                    LIMIT 81;""")
        data = cur.fetchall()
        cur.execute("""SELECT count(id) FROM analytic_project_code;""")
        extra_datas[2] = cur.fetchall()[0][0]
        extra_datas[0] = "Project List"
    elif what == 'fuel-price':
        data = {}
        extra_datas[0] = "Fuel Prices"
        cur.execute("SELECT * FROM fuel_price_history ORDER BY end_date DESC LIMIT 81;")
        data = cur.fetchall()
        cur.execute("SELECT count(id) FROM fuel_price_history;")
        extra_datas[2] = cur.fetchone()[0]
    elif what in ('type','class','unit','capacity','brand','owner'):
        query_table_dct = {'type':'machine_type','class':'machine_class','brand':'fleet_vehicle_model_brand',
                           'unit':'res_company','capacity':'vehicle_machine_config','owner':'vehicle_owner'}
        table_name_dct = {'type':'Machine Type','class':'Machine Class','brand':'Vehicle Brand',
                          'unit':'Business Unit','capacity':'Machine Capacity','owner':'Vehicle Owner'}
        cur.execute(f"SELECT id,name FROM {query_table_dct[what]} WHERE name iLIKE '%{search_wildcard}%' ORDER BY CASE WHEN name = '{search_wildcard}' THEN 0 ELSE 1 END;")
        data = cur.fetchall()
        cur.execute(f"SELECT count(id) FROM {query_table_dct[what]};")
        extra_datas[2] = cur.fetchone()[0]
        extra_datas[0] = table_name_dct[what]
    elif what == 'project-stat':
        cur.execute(f""" SELECT pj.id , pj.name , pj.code , emp.name, stat.location , stat.pj_start_date, estimate_feet, will_sud , estimate_sud, estimate_duty , estimate_fuel , estimate_expense , estimate_day
                    FROM project_statistics stat 
                        INNER JOIN analytic_project_code pj ON stat.project_id = pj.id 
                    INNER JOIN employee emp ON emp.id = stat.supervisor_id
                        {where_filter_clause[:-6]}
                    ORDER BY pj.name LIMIT 81;""")
        
        data = cur.fetchall()
        cur.execute("""SELECT count(id) FROM project_statistics;""")
        extra_datas[2] = cur.fetchall()[0][0]
        extra_datas[0] = "Project Statistics"
    elif what == 'employee':
        extra_datas[0] = 'Employee'
        cur.execute("SELECT emp.id,'MD-' || code,emp.name,emp_gp.name FROM employee emp LEFT JOIN employee_group emp_gp ON emp.employee_group_id = emp_gp.id;")
        data = cur.fetchall()
        cur.execute("SELECT id,name FROM employee_group;")
        extra_datas.append(cur.fetchall())
        cur.execute("SELECT count(id) FROM employee;")
        extra_datas[2] = cur.fetchone()[0]
    elif what == 'employee-group':
        extra_datas[0] = 'Employee Group'
        cur.execute("SELECT id,name FROM employee_group;")
        data = cur.fetchall()
        cur.execute("SELECT count(id) FROM employee_group;")
        extra_datas[2] = cur.fetchone()[0]
    elif what == 'ajt' or what == 'ajf':
        db = {'ajt':'activity_job_type','ajf':'activity_job_function'}
        naming = {'ajt':'Activity Job Type','ajf':'Activity Job Function'}
        extra_datas[0] = naming[what]
        cur.execute(f"SELECT id,name FROM {db[what]};")
        data = cur.fetchall()
        cur.execute(f"SELECT count(id) FROM {db[what]};")
        extra_datas[2] = cur.fetchone()[0]
        
    cur.close()
    conn.close()

    return render_template('configurations.html',mgs=mgs,datas=data,extra_datas = extra_datas,what=what,project_datas=project_datas, current_role = role)
        
def get_necessary_data_for_imports():
    global mappings,machine_brand_mapping,machine_capacity_mapping,machine_class_mapping,machine_owner_mapping,machine_type_mapping,business_unit_mapping
    
    conn = db_connect()
    cur = conn.cursor()
    
    cur.execute("SELECT id,name FROM machine_type")
    datas = cur.fetchall()
    machine_type_mapping = {item[1]: item[0] for item in datas}

    cur.execute("SELECT id,name FROM machine_class")
    datas = cur.fetchall()
    machine_class_mapping = {item[1]: item[0] for item in datas}

    cur.execute("SELECT id,name FROM res_company")
    datas = cur.fetchall()
    business_unit_mapping = {item[1]: item[0] for item in datas}

    cur.execute("SELECT id,name FROM vehicle_machine_config")
    datas = cur.fetchall()
    machine_capacity_mapping = {item[1]: item[0] for item in datas}

    cur.execute("SELECT id,name FROM fleet_vehicle_model_brand")
    datas = cur.fetchall()
    machine_brand_mapping = {item[1]: item[0] for item in datas}

    cur.execute("SELECT id,name FROM vehicle_owner")
    datas = cur.fetchall()
    machine_owner_mapping = {item[1]: item[0] for item in datas}

    mappings = [business_unit_mapping,machine_type_mapping,machine_capacity_mapping,machine_brand_mapping,machine_owner_mapping,machine_class_mapping]


@views.route("/upload-machine-details",methods=['GET','POST'])
def upload_machine_details():

    if "cpu_user_id" not in session:
        return redirect(url_for("auth.authenticate"))

    mgs = None
    conn = db_connect()
    cur = conn.cursor()

    user_id = session["cpu_user_id"]
    cur.execute("SELECT user_access_id FROM user_auth WHERE id = %s;",(user_id,))
    role = cur.fetchone()[0]

    if request.method == 'POST':
        upload_file = request.files["upload_excel_machine_details"]
        excel_file_type = request.form.get('selectedOption')
        what_dct = {"Machine List":"machine","Project Code":"project","Machine Details":"unit"}
        if upload_file.filename != '' and upload_file.filename.endswith(".xlsx"):
            workbook = load_workbook(filename=upload_file,data_only=True,read_only=True)
            # Select the worksheet to read from
            try:
                worksheet = workbook[excel_file_type]  # Replace 'Sheet1' with the actual sheet name
            except:
                mgs = f"The sheet name of your Imported Excel file doesn't match with <strong>{excel_file_type}<&#47;strong>"
                return redirect(url_for('views.configurations',what=what_dct.get(excel_file_type),mgs=mgs))
            # GET
            get_necessary_data_for_imports()
            # Iterate over rows in the worksheet

            if excel_file_type == "Machine Details":
                query_lst = [
                    """INSERT INTO res_company (name) VALUES """,
                    """INSERT INTO machine_type (name) VALUES """,
                    """INSERT INTO vehicle_machine_config (name) VALUES """,
                    """INSERT INTO fleet_vehicle_model_brand (name) VALUES """,
                    """INSERT INTO vehicle_owner (name) VALUES """,
                    """INSERT INTO machine_class (name) VALUES """
                ]
                for row in worksheet.iter_rows(min_row=2):  # Start from the second row (adjust as needed)
                    # Access data for each cell in the row
                    for idx,row_data in enumerate(row):
                        query_lst[idx] += f"""('{row_data.value}'),""" if row_data.value and row_data.value.strip() != "" else ""
                queries = [each_query[:-1] for each_query in query_lst if not each_query.strip().endswith("VALUES")]
                mgs = catch_db_insert_error(cur,conn,queries)
            elif excel_file_type == "Machine List":
                machine_list_insert_query = """INSERT INTO fleet_vehicle 
                (business_unit_id,machine_type_id,machine_config_id,brand_id,owner_name_id,machine_class_id,machine_name) VALUES """
                for row_counter , row in enumerate(worksheet.iter_rows(min_row=2),start=1):  # Start from the second row (adjust as needed)        
                    stng = "("
                    for idx,mapping in enumerate(mappings):
                        idd = mapping.get(row[idx].value)
                        if idd is None:
                            mgs = f"Unknown Field or Unknow Value at <strong class='text-danger'>  Row : ({row_counter}) {row[idx].value} <&#47;strong> "
                            return redirect(url_for('views.configurations',what='machine',mgs=mgs))
                        else:
                            stng += f"{idd},"
                    machine_list_insert_query += stng + f"'{row[6].value.strip()}'),"
                mgs = catch_db_insert_error(cur,conn,[machine_list_insert_query[:-1]+ 'ON CONFLICT (machine_name) DO NOTHING;']) 
            elif excel_file_type == 'Work Done':
                machine_list_insert_query = """ INSERT INTO  project_each_day_work_done 
                (project_id,duty_date,work_done,constraint_unique_text) VALUES """
                for row_counter, row in enumerate(worksheet.iter_rows(min_row=2),start=1):
                    if not row[0].value or not row[1].value or row[2].value is None:
                        return redirect(url_for('views.machine_details',mgs=f"<strong>Blank Field at ROW No. {row_counter}<&#47;strong>"))
                    else:
                        cur.execute("SELECT id FROM analytic_project_code WHERE code = %s",(row[1].value,))
                        id = cur.fetchall()
                        if not id:
                            return redirect(url_for('views.machine_details',mgs=f"<strong>Project Code {row[1].value} doesn't exit!!<&#47;strong>"))   
                        unique_constraint_for_work_done = row[0].value.strftime('%Y-%m-%d') + "|" + str(id[0][0])
                        machine_list_insert_query += f"('{id[0][0]}','{row[0].value}','{row[2].value}','{unique_constraint_for_work_done}'),"
                mgs = catch_db_insert_error(cur,conn,[machine_list_insert_query[:-1] + 'ON CONFLICT (constraint_unique_text) DO NOTHING;'])                
            elif excel_file_type == "Project Code":
                machine_list_insert_query = ''' INSERT INTO analytic_project_code
                (code,pj_group,name,type,business_unit_id) VALUES '''
                for row_counter , row in enumerate(worksheet.iter_rows(min_row=2),start=1):
                    if not row[0].value  or not row[1].value  or  not row[2].value or not row[3].value or not row[4].value:
                        mgs = f"<strong>Blank Field at ROW No. {row_counter}<&#47;strong>"
                        return redirect(url_for('views.machine_details',mgs=mgs))
                    else:
                        unit = mappings[0].get(row[4].value)
                        if not unit:
                            return redirect(url_for('views.machine_details',mgs=f"<strong>Invalid Buinsess Unit <{row[4].value}> at ROW No. {row_counter}<&#47;strong>"))
                        machine_list_insert_query += f"('{row[0].value}','{row[1].value}','{row[2].value}','{row[3].value}','{unit}'),"
                machine_list_insert_query = machine_list_insert_query[:-1] + 'ON CONFLICT (code) DO NOTHING;'
                mgs = catch_db_insert_error(cur,conn,[machine_list_insert_query]) 

            workbook.close()
            cur.close()
            conn.close()

            return redirect(url_for('views.configurations',what=what_dct.get(excel_file_type),mgs=mgs))
        else:    
            return redirect(url_for('views.configurations',what=what_dct.get(excel_file_type,"project"),mgs="Invalid File Type"))
    else:
        return render_template("home.html", current_role = role)

@views.route("/get-pj-datas")
def get_pj_datas():
    conn = db_connect()
    cur = conn.cursor()

    cur.execute("SELECT id ,name , code FROM analytic_project_code")
    pj_datas = cur.fetchall()
    
    return jsonify(pj_datas)


@views.route("upload-duty",methods=['GET','POST'])
def upload_duty():
    mgs = None
    if request.method == 'POST':
        upload_file = request.files['upload_duty_data']
        month = request.form.get("selectedMonth")
        year = request.form.get("selectedYear")
        query = request.form.get('selectedQuery')
        what_dct = {"Duty Query":'duty',"Expenses Query":'expense','Services Query':"service"}
        if upload_file.filename != '' and upload_file.filename.endswith(".xlsx"):
            workbook = load_workbook(filename=upload_file,read_only=True,data_only=True)
            # Select the worksheet to read from
            try:
                worksheet = workbook[query]  # Replace 'Sheet1' with the actual sheet name
            except:
                return redirect(url_for('views.show_transactions',what=what_dct.get(query,'duty'),mgs=f"Sheet Name must be  {query}"))     
            # Iterate over rows in the worksheet
            conn = db_connect()
            cur = conn.cursor()

            if query == 'Duty Query':
                insert_statement_query = """ INSERT INTO duty_odoo_report 
                (project_id,duty_date,machine_id,operator_name,morg_start,morg_end,aftn_start,aftn_end,evn_start,evn_end,total_hr,totaluse_fuel,hrper_rate,fuel_price,duty_amt,fuel_amt,way,complete_feet,complete_sud)  VALUES """
                for row_counter , row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    # 3,4,10,11,14,15,16,17,18,19,20,30,33,37,38,43,46
                    # 3,4,10,11,27,28,29,39,31,32,33,43,46,52,53,56,59
                    if row[3].value is None or row[4].value is None or row[10].value is None or row[11].value is None or row[27].value is None or row[28].value is None or row[29].value is None or row[30].value is None or row[31].value is None or row[32].value is None or row[33].value is None or row[46].value is None or row[50].value is None or row[51].value is None:
                        return redirect(url_for('views.show_transactions',what=what_dct.get(query,'duty'),mgs=f"Invalid Field or Blank Field at Row - {row_counter}"))     
                    else:
                        if year != 'all':
                            if row[4].value.month != int(month) and row[4].value.year != int(year):
                                return redirect(url_for('views.duty_query',mgs=f"Selected date doen't match with Imported date at Row - {row_counter}"))
                        com_feet = row[43].value if row[43].value else 0.00
                        way = row[56].value if row[56].value else 0
                        com_sud = row[59].value if row[59].value else 0.00
                        cur.execute("SELECT id FROM analytic_project_code WHERE code = %s",(row[3].value,))
                        pj_id = cur.fetchone()
                        pj_id = pj_id[0] if pj_id else pj_id
                        cur.execute("SELECT id FROM fleet_vehicle WHERE machine_name = %s",(row[10].value,))
                        machine_id = cur.fetchone()
                        machine_id = machine_id[0] if machine_id else machine_id
                        if not machine_id or not pj_id:
                            mgs = f"Invalid field or Unknown Field at Row - {row_counter} <br>"
                            mgs = mgs + f"Project Code - {row[3].value.replace('/','&#47;')} doesnt't match with system <br>" if not pj_id else mgs
                            mgs = mgs + f"Machine Name - {row[10].value.replace('/','&#47;')} doesnt't match with system <br>" if not machine_id else mgs
                            return redirect(url_for('views.show_transactions',what=what_dct.get(query,'duty'),mgs=mgs))
                        insert_statement_query += f""" ({pj_id},'{row[4].value.strftime("%Y-%m-%d")}',{machine_id},'{row[11].value}','{row[27].value}','{row[28].value}','{row[29].value}','{row[30].value}','{row[31].value}','{row[32].value}','{row[33].value}','{row[46].value}','{row[50].value}','{row[51].value}',{row[52].value},{row[53].value},'{way}','{com_feet}','{com_sud}'),"""
                table = 'duty_odoo_report'
            elif query == 'Expenses Query':
                insert_statement_query = """ INSERT INTO expense_prepaid 
                (res_company_id,project_id,duty_date,expense_amt) VALUES """
                cur.execute("SELECT name,id from res_company;")
                unit_ids = cur.fetchall()
                unit_ids_dct = {data[0]:data[1] for data in unit_ids}
                for row_counter , row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    if row[5].value is None or row[8].value is None or row[10].value is None:
                        return redirect(url_for('views.duty_query',mgs=f"Invalid field or Blank field at Row - {row_counter}"))
                    if year != 'all':
                        if row[10].value.month != int(month) and row[10].value.year != int(year):
                            return redirect(url_for('views.duty_query',mgs=f"Selected date doen't match with Imported date at Row - {row_counter}"))
                    unit_id = unit_ids_dct.get(row[5].value.strip())
                    cur.execute("SELECT id FROM analytic_project_code WHERE code = %s",(row[8].value.strip(),))
                    id = cur.fetchall()
                    if unit_id and id != []:
                        if row[38].value.strip() == 'P&L':
                            expenses = row[22].value if row[22].value else 0
                            insert_statement_query += f""" ('{unit_id}','{id[0][0]}','{row[10].value}',{expenses}),"""
                    else:
                        return redirect(url_for('views.show_transactions',what=what_dct.get(query,'duty'),mgs=f"Unknown Project Code or Unknown Unit at Row - {row_counter}"))                        
                table = "expense_prepaid"
            elif query == "Services Query":
                insert_statement_query = """ INSERT INTO service_datas (service_date,business_unit_id,machine_id,
                category_id,parts_price,parts_qty) VALUES """
                cur.execute("SELECT machine_name,id FROM fleet_vehicle;")
                vehicle_datas = {data[0]:data[1] for data in cur.fetchall()}
                cur.execute("SELECT name,id FROM res_company;")
                bi_datas = {data[0]:data[1] for data in cur.fetchall()}
                cur.execute("SELECT name,id FROM service_category;")
                category_datas = {data[0]:data[1] for data in cur.fetchall()}
                for row_counter , row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    # print(row[0].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value)
                    if None in (row[0].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value) or row[2].value not in bi_datas or row[3].value not in vehicle_datas:
                        print("Blank Field")
                    category_id = row[4].value.strip()
                    if not category_datas.get(category_id):
                        cur.execute("INSERT INTO service_category(name) VALUES (%s) ON CONFLICT (name) DO NOTHING RETURNING id;",(category_id,))
                        category_id = cur.fetchall()[0][0]
                    insert_statement_query += f""" ('{row[0].value}','{bi_datas.get(row[2].value.strip())}','{vehicle_datas.get(row[3].value.strip())}','{category_id}','{row[5].value.strip()}','{row[6].value.strip()}') """
            # print(insert_statement_query)
            if year != 'all':
                s_date , e_date = get_first_and_last_day(int(month),int(year))
                cur.execute("""DELETE FROM {} WHERE duty_date BETWEEN '{}' AND '{}' """.format(table,s_date,e_date))
                conn.commit()
            mgs = catch_db_insert_error(cur,conn,[insert_statement_query[:-1]])
            cur.close()
            conn.close()
        else:
            mgs = "Invalid File Type.."
    return redirect(url_for('views.show_transactions',what=what_dct.get(query,'duty'),mgs=mgs))

@views.route("/delete-data/<db>/<id>")
def delete_data(db,id):
    conn = db_connect()
    cur = conn.cursor()
    mgs = 'Success'
    try:
        if db == 'income_expense':
            cur.execute("DELETE FROM income_expense_line WHERE income_expense_id = %s;",(id,))
        cur.execute(f'DELETE FROM {db} WHERE id = {id}')
        conn.commit()
    except IntegrityError as err:
        print(err)
        mgs = err
        conn.rollback()
    return mgs    


@views.route("/edit-data/<db>/<id>/<stng>")
def edit_data_in_db(db,id,stng:str):
    queries_values = { 'project_statistics' : """ UPDATE project_statistics SET supervisior = %s,estimate_feet = %s,
                                                    will_sud = %s,estimate_sud = %s,estimate_duty = %s,estimate_fuel = %s,
                                                    estimate_expense = %s WHERE id = %s;""",
                        'fleet_vehicle' : """ UPDATE fleet_vehicle SET business_unit_id = %s,machine_type_id = %s,
                                                machine_config_id = %s,brand_id = %s,owner_name_id = %s,
                                                machine_class_id = %s,machine_name = %s WHERE id = %s;""",
                        'analytic_project_code' : """ UPDATE analytic_project_code SET code = %s,name = %s,pj_group = %s,
                                                type = %s,business_unit_id = %s , finished_state = %s WHERE id = %s"""
    }
    datas = stng.split(",")
    lst = []
    if db == 'fleet_vehicle':
        "business_unit_id,machine_type_id,machine_config_id,brand_id,owner_name_id,machine_class_id,machine_name"
        lst = [datas[3],datas[1],datas[4],datas[5],datas[6],datas[2]]
        for idx,mapping in enumerate(mappings):
            dt = mapping.get(lst[idx].replace('thisIsSlash','/'))
            if dt is None:
                return f"""["Error","{lst[idx]}"]"""
            lst[idx] = dt
        lst.append(datas[0].replace('thisIsSlash','/'))
        lst.append(int(id))
    elif db == 'analytic_project_code':
        lst = list(datas)
        conn = db_connect()
        cur = conn.cursor()
        cur.execute(f"SELECT id FROM res_company WHERE name = '{lst[4].replace('thisIsSlash','/')}';")
        unit = cur.fetchall()
        if lst[5] not in ['True','False'] or unit == []:
            return f"""["Error","{lst[5]} | {lst[4]}"]"""
        lst[4] = str(unit[0][0]) 
        # print(lst)
        lst = [dt.replace('thisIsSlash','/') for dt in lst]

    inserted_values_query = {'project_statistics':tuple(datas)+(id,),
                            'fleet_vehicle':tuple(lst),
                            'analytic_project_code':tuple(lst)+(id,)}
    conn = db_connect()
    cur = conn.cursor()
    # print(queries_values[db],inserted_values_query[db])
    try:
        cur.execute(queries_values[db],inserted_values_query[db])
        conn.commit()
    except IntegrityError as err:
        print(err)
        conn.rollback()
    cur.close()
    conn.close()
    return """["Success"]"""

@views.route("/offset-display/<for_what>/<ofset>")
def offset_display(for_what,ofset):
    conn = db_connect()
    cur = conn.cursor()

    if "cpu_user_id" not in session:
        return redirect(url_for("auth.authenticate"))

    user_id = session["cpu_user_id"]
    cur.execute("SELECT user_access_id FROM user_auth WHERE id = %s;",(user_id,))
    role = cur.fetchone()[0]
    if role in [3,4]:
        project_filter_query = ''
        income_expense_project = ''
    else:
        project_filter_query = f"""
            WHERE das.pj_id in (SELECT project_id FROM project_user_access WHERE user_id = '{user_id}')
        """
        income_expense_project = f" WHERE form.project_id in (SELECT project_id FROM project_user_access WHERE user_id = '{user_id}') "

    queries_dct = {
        "pj-data-changeable" : f""" SELECT 
                    code,pj.name,p_group.name,p_type.name,bi.name,pj.id
                    FROM analytic_project_code pj 
                    LEFT JOIN res_company bi
                    ON pj.business_unit_id = bi.id
                    LEFT JOIN project_group AS p_group
                    ON p_group.id = pj.pj_group_id
                    LEFT JOIN project_type AS p_type
                    ON p_type.id = pj.pj_type_id
                    LIMIT 81 OFFSET {ofset};""",
        "duty-query-changeable": f""" SELECT 
                    dty.duty_date,pj.code,pj.name,fv.machine_name,dty.operator_name,dty.morg_start,dty.morg_end,
                    dty.aftn_start,dty.aftn_end,dty.evn_start,dty.evn_end,dty.total_hr,dty.hrper_rate,
                    dty.totaluse_fuel,dty.fuel_price,dty.duty_amt,dty.fuel_amt,dty.total_amt,dty.way,
                    dty.complete_feet, dty.complete_sud 
                    FROM duty_odoo_report dty
                    LEFT JOIN fleet_vehicle fv ON fv.id = dty.machine_id
                    LEFT JOIN analytic_project_code pj ON pj.id = dty.project_id ORDER BY dty.duty_date ASC 
                    LIMIT 81 OFFSET {ofset}""",
        "machine-list-changeable" : f"""
                        SELECT pj.name,fv.machine_name,mt.name,mc.name,rc.name,mcf.name,vb.name,vo.name,fv.id
                        FROM fleet_vehicle fv
                        LEFT JOIN machine_type mt ON mt.id = fv.machine_type_id 
                        LEFT JOIN machine_class mc ON mc.id = fv.machine_class_id
                        LEFT JOIN res_company rc ON rc.id = fv.business_unit_id
                        LEFT JOIN fleet_vehicle_model_brand vb ON vb.id = fv.brand_id
                        LEFT JOIN vehicle_owner vo ON vo.id = fv.owner_name_id
                        LEFT JOIN vehicle_machine_config mcf ON mcf.id = fv.machine_config_id 
                        LEFT JOIN 
                            ( SELECT project_id,machine_id FROM machines_history 
                                WHERE CURRENT_TIMESTAMP BETWEEN start_time AND COALESCE(end_time,CURRENT_TIMESTAMP)
                            ) AS history ON history.machine_id = fv.id
                        LEFT JOIN analytic_project_code AS pj ON pj.id = history.project_id
                        LIMIT 81 OFFSET {ofset}""",
        "expense-list-changeable" : f"""  SELECT
                            exp.duty_date,pj.code,pj.name,bi.name,exp.expense_amt
                            FROM expense_prepaid AS exp
                            INNER JOIN analytic_project_code AS pj
                            ON pj.id = exp.project_id
                            INNER JOIN res_company bi
                            ON bi.id = exp.res_company_id
                            LIMIT 81 OFFSET {ofset};""",
        "income-expense-changeable":f"""SELECT form.set_date,pj.code,pj.name,form.income_expense_no,
                                        CASE WHEN form.income_status = 't' THEN 'INCOME' ELSE 'EXPENSE' END,
                                        SUM(line.amt),form.id
                                        FROM income_expense AS form 
                                        INNER JOIN income_expense_line AS line 
                                        ON line.income_expense_id = form.id
                                        INNER JOIN analytic_project_code AS pj
                                        ON pj.id = form.project_id
                                        {income_expense_project}
                                        GROUP BY form.id,pj.code,pj.name
                                        ORDER BY MAX(pj.id),form.set_date DESC
                                        LIMIT 81 OFFSET {ofset};""",
        "daily-activity-changeable":f""" WITH daily_activity_summary AS (
                            SELECT DISTINCT ON (da.id)
                                da.id,
                                da.set_date,
                                pj.code,
                                pj.name,
                                da.daily_activity_no,
                                CASE WHEN da.working_status = 't' THEN 'Working' ELSE 'Not Working' END AS working_status,
                                da.remark_for_not_working,
                                da.wealther_effect_percent,
                                da.complete_feet,
                                da.complete_sud,
                                pj.id AS pj_id
                            FROM daily_activity da
                            JOIN analytic_project_code pj ON da.project_id = pj.id
                            LEFT JOIN daily_activity_accident_lines AS accident ON accident.daily_activity_id = da.id
                        ),
                        duty_hours_summary AS (
                        
                            SELECT daily_activity_id, EXTRACT(HOUR FROM sum(duty_hour)) || ':' || EXTRACT(MINUTE FROM sum(duty_hour)) AS total_duty_hour, SUM(used_fuel)  AS total_used_fuel, SUM(way) AS ways
                            FROM daily_activity_lines
                            GROUP BY daily_activity_id
                        ),
                        present_people_summary AS (
                            SELECT daily_activity_id, SUM(present_people) AS total_present_people
                            FROM employee_group_project_line
                            GROUP BY daily_activity_id
                        ),
                        present_machines_summary AS (
                            SELECT daily_activity_id, SUM(present_machines) AS total_present_machines
                            FROM machines_history_project
                            GROUP BY daily_activity_id
                        ),
                        accident_summary AS (
                            SELECT daily_activity_id, COUNT(machine_id) AS machine_count, bool_or(accident_status) AS any_accident
                            FROM daily_activity_accident_lines
                            GROUP BY daily_activity_id
                        )
                        SELECT
                            das.set_date,
                            das.code,
                            das.name,
                            das.daily_activity_no,
                            das.working_status,
                            das.remark_for_not_working,
                            das.wealther_effect_percent,
                            das.complete_feet,
                            das.complete_sud,
                            COALESCE(dh.total_duty_hour,'0:00:00'),
                            COALESCE(dh.total_used_fuel,'0.0'),
                            pp.total_present_people,
                            pm.total_present_machines,
                            COALESCE(a.machine_count,0),
                            COALESCE(a.any_accident,'FALSE'),
                            COALESCE(dh.ways,'0.0'),
                            das.id
                        FROM daily_activity_summary das
                        LEFT JOIN duty_hours_summary dh ON das.id = dh.daily_activity_id
                        LEFT JOIN present_people_summary pp ON das.id = pp.daily_activity_id
                        LEFT JOIN present_machines_summary pm ON das.id = pm.daily_activity_id
                        LEFT JOIN accident_summary a ON das.id = a.daily_activity_id
                        {project_filter_query}
                        ORDER BY das.pj_id,das.set_date DESC,das.daily_activity_no DESC
                        LIMIT 81 OFFSET {ofset};"""
    }
    print(queries_dct[for_what])
    cur.execute(queries_dct[for_what])
    result_datas = cur.fetchall()
    if for_what == 'dty':
        result_datas = [
            (str(d[0]), d[1], d[2], d[3], d[4], str(d[5]), str(d[6]), str(d[7]), str(d[8]), str(d[9]), str(d[10]),
            str(d[11]), str(d[12]), str(d[13]), str(d[14]), str(d[15]), str(d[16]), str(d[17]), d[18], str(d[19]), 
            str(d[20])) for d in result_datas
            ]
    elif for_what == 'income-expense-changeable' or for_what == 'daily-activity-changeable':
        for idx,data in enumerate(result_datas):
            data = list(data)
            data[0] = str(data[0])
            result_datas[idx] = data
    return jsonify(result_datas)


@views.route("/get-api/<for_what>/<data>")
def call_api(for_what,data:str):
    conn = db_connect()
    cur = conn.cursor()
    data = data.replace("thisIsSlash","/")
    if for_what == 'project-stat':
        cur.execute("""  
        SELECT
            pj.id,emp.name,pj.name,stat.location,stat.pj_start_date,sup.name
        FROM project_statistics AS stat
            INNER JOIN analytic_project_code AS pj
            ON stat.project_id = pj.id
            INNER JOIN employee AS emp
            ON stat.ho_acc_id = emp.id
            INNER JOIN employee AS sup
            ON stat.supervisor_id = sup.id 
        WHERE pj.code = %s or pj.name = %s;""",(data,data))
    elif for_what == 'call-duty-price-for-each-machine':
        cur.execute("SELECT duty_price FROM duty_price_history WHERE machine_id = %s AND end_date IS NULL;",(data,))
        price = cur.fetchone()
        return jsonify(price)
    elif for_what == 'project-check-for-stats':
        cur.execute("SELECT pj.id FROM analytic_project_code AS pj LEFT JOIN project_statistics AS stat ON pj.id = stat.project_id WHERE stat.project_id is NULL AND (name = %s or code = %s) LIMIT 1;",(data,data))
    elif for_what == 'project-check-for-transfer-stats':
        cur.execute("SELECT id FROM analytic_project_code WHERE code = %s or name = %s;",(data,data))
    elif for_what == 'machine-check':
        cur.execute(""" 
                        SELECT
                            car.id AS machine_id,
                            car.machine_name,
                            tp.name AS machine_type
                        FROM
                            fleet_vehicle AS car
                        LEFT JOIN machines_history AS history
                        ON car.id = history.machine_id
                        LEFT JOIN machine_type AS tp
                        ON tp.id = car.machine_type_id
                        WHERE history.project_id = 4 AND history.end_time IS NULL AND car.machine_name = %s;
            """,(data,))
    elif for_what == 'employee-group-check':
        cur.execute("SELECT id,name FROM employee_group WHERE name = %s;",(data,))
    elif for_what == 'employee-group-edit':
        edit_id , edit_amt = data.split("~~")
        cur.execute("UPDATE employee_group_project SET assigned_people = %s WHERE id = %s RETURNING id;",(edit_amt,edit_id))
        conn.commit()
    elif for_what == 'duty-amt-check':
        machine_id, set_date = data.split("~~")
        set_date = datetime.strptime(set_date, "%d-%m-%Y").strftime("%Y-%m-%d")
        cur.execute(""" SELECT id FROM duty_price_history  WHERE machine_id = %s AND %s BETWEEN start_date AND COALESCE(end_date,%s) LIMIT 1;""",(machine_id,set_date,set_date))
    elif for_what == 'accountant-supervisor-check':
        sup , acc = data.split("~|~")
        result = []
        cur.execute("SELECT emp.id FROM employee emp INNER JOIN employee_group emp_gp ON emp.employee_group_id = emp_gp.id WHERE emp_gp.name = 'SUPERVISOR' AND LOWER(emp.name) = %s LIMIT 1;",(sup.lower(),))
        result.append(cur.fetchone())
        cur.execute("SELECT emp.id FROM employee emp INNER JOIN employee_group emp_gp ON emp.employee_group_id = emp_gp.id WHERE emp_gp.name = 'ACCOUNTANT' AND LOWER(emp.name) = %s LIMIT 1;",(acc.lower(),))
        result.append(cur.fetchone())
        return jsonify(result)
    elif for_what == 'delete-machines-history':
        cur.execute("UPDATE machines_history SET end_time = NOW() WHERE machine_id =  %s AND end_time IS NULL;",(data,))
        cur.execute("INSERT INTO machines_history (project_id,machine_id) VALUES(4,%s);",(data,))
        conn.commit()
        cur.execute("SELECT 1;")
    elif for_what == 'delete-employee-group-history':
        cur.execute("DELETE FROM employee_group_project WHERE id = %s RETURNING id;",(data,))
        conn.commit()
    elif for_what == 'check-machines-within-date':
        pj_id, set_date = data.split("~~") 
        cur.execute(""" 
            SELECT car.machine_name,car.id FROM machines_history
                LEFT JOIN fleet_vehicle AS car ON car.id = machines_history.machine_id
            WHERE project_id = %s AND %s BETWEEN start_time::date AND COALESCE(end_time::date,CURRENT_DATE);
        """,(pj_id,datetime.strptime(set_date, "%d-%m-%Y").strftime("%Y-%m-%d")))
    elif for_what == 'transfer_machine_project':
        # print(data.split("~~"))
        his_id , project_id , start_time , machine_id = data.split("~~")
        try:
            cur.execute("UPDATE machines_history SET end_time = %s WHERE id = %s;",(start_time,his_id))
            cur.execute("INSERT INTO machines_history (project_id,start_time,machine_id) VALUES (%s,%s,%s);",(project_id,start_time,machine_id))
            if project_id == '4':
                end_date = datetime.fromisoformat(start_time).date().strftime('%Y-%m-%d')
                cur.execute("UPDATE duty_price_history SET end_date = %s WHERE machine_id = %s AND end_date IS NULL AND start_date <= %s RETURNING id;",(end_date,machine_id,end_date))
            conn.commit()
        except IntegrityError as err:
            print(err)
        return ""
    result = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(result)
